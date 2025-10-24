"""Excel document generator using openpyxl."""

from pathlib import Path
from typing import Dict, Any
import re
from openpyxl import load_workbook
from .base import DocumentGenerator


class ExcelGenerator(DocumentGenerator):
    """Generator for Microsoft Excel documents."""
    
    def generate(self, template_path: Path, fields: Dict[str, Any], output_path: Path) -> Path:
        """
        Generate an Excel document from a template by replacing placeholders.
        
        Placeholders in the template should be in the format {{field_name}}.
        """
        # Load the template
        wb = load_workbook(template_path)
        
        # Iterate through all worksheets
        for sheet in wb.worksheets:
            # Iterate through all cells
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Replace placeholders
                        cell_value = cell.value
                        for field_name, field_value in fields.items():
                            placeholder = f"{{{{{field_name}}}}}"
                            if placeholder in cell_value:
                                cell_value = cell_value.replace(placeholder, str(field_value))
                        cell.value = cell_value
        
        # Save the workbook
        wb.save(output_path)
        return output_path
    
    def get_template_fields(self, template_path: Path) -> list[str]:
        """Extract field names from an Excel template."""
        wb = load_workbook(template_path)
        fields = set()
        
        # Find placeholders in format {{field_name}}
        pattern = r'\{\{([^}]+)\}\}'
        
        # Search in all worksheets
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        matches = re.findall(pattern, cell.value)
                        fields.update(matches)
        
        return sorted(list(fields))
